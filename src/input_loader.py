"""Load the list of X/Twitter post links to capture.

Two sources:
    <path>.xlsx      an Excel sheet   -> load_excel
    "-" / "paste"    pasted on stdin  -> load_stdin (one link per line)

(A non-Excel file passed as <path> is tolerated: load_excel falls back to
reading it as a plain link list.)

Every source yields the same row shape:
    {"category", "account_name", "link", "post_link", "platform"}

Recognized layouts:
  * A header row naming a link column ("link" / "url" / "post link" / "tweet")
    is read as a table; optional "account"/"name"/"handle" and "category"
    columns are picked up when present.
  * Otherwise it's read as a plain list: any cell/line containing an http(s)
    URL is a link; a non-URL line becomes the current category header.

Only X/Twitter links are kept — this project is X-only.
"""
import csv
import io
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from platforms import platform_of  # noqa: E402

_URL_RE = re.compile(r"https?://\S+", re.I)
_LINK_HEADERS = ("link", "url", "post link", "post_link", "tweet", "tweet link",
                 "tweet url")
_ACCOUNT_HEADERS = ("account", "account name", "account_name", "name", "handle",
                    "page name", "author")
_CATEGORY_HEADERS = ("category", "section", "group")
_IGNORE_HEADERS = ("s.no", "serial", "sr", "#", "no", "followers", "views")


def is_x_url(url: str) -> bool:
    u = (url or "").lower()
    return "x.com" in u or "twitter.com" in u


def derive_name(url: str) -> str:
    """A display name from an X URL: '@handle' when present, else 'X post'.
    /i/status/ links hide the handle — the capture reads it from the page."""
    m = re.search(r"(?:x\.com|twitter\.com)/([^/?#]+)/status", (url or "").lower())
    if m and m.group(1) not in ("i", "intent", "home"):
        return "@" + m.group(1)
    return "X post"


def _clean_url(link: str) -> str:
    """Trim whitespace and trailing punctuation a URL never really ends in
    (e.g. a stray '.' or ',' pasted from a sheet: '.../374.' -> '.../374')."""
    link = (link or "").strip()
    return link.rstrip(".,;:!?)]}>\"'")


def _row(category, account, link):
    link = _clean_url(link)
    return {
        "category": (category or "Uncategorized").strip() or "Uncategorized",
        "account_name": (account or "").strip() or derive_name(link),
        "link": link,
        "post_link": link,
        "platform": platform_of(link),
    }


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #
def load_excel(path: str) -> list:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError as e:
        raise RuntimeError(
            "Reading .xlsx needs openpyxl — run `python install.py` "
            "(or `pip install openpyxl`).") from e

    import zipfile
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except (zipfile.BadZipFile, InvalidFileException):
        # Named .xlsx but it's really a plain-text / CSV list (a common mistake
        # when a text file is just renamed). Fall back to the list reader.
        print(f"[input] {Path(path).name} isn't a real Excel file — "
              "reading it as a plain link list instead")
        return load_delimited(path)
    ws = wb.active
    grid = [[("" if c is None else str(c)).strip() for c in r]
            for r in ws.iter_rows(values_only=True)]
    wb.close()
    return _rows_from_grid(grid)


# --------------------------------------------------------------------------- #
# CSV / pasted list / stdin
# --------------------------------------------------------------------------- #
def load_delimited(path_or_text: str, is_text: bool = False) -> list:
    raw = path_or_text if is_text else Path(path_or_text).read_text()
    # A .csv (or any comma-separated paste) parses cleanly as CSV; a plain
    # one-link-per-line paste parses as single-column CSV, which is fine.
    grid = [[c.strip() for c in r] for r in csv.reader(io.StringIO(raw))]
    return _rows_from_grid(grid)


def load_stdin() -> list:
    print("[input] paste X/Twitter links (one per line), then Ctrl-D:")
    return load_delimited(sys.stdin.read(), is_text=True)


# --------------------------------------------------------------------------- #
# Shared grid parser (table layout OR plain list)
# --------------------------------------------------------------------------- #
def _header_index(grid):
    """Row index + column map if a header row names a link column, else None."""
    for i, cells in enumerate(grid[:5]):
        low = [c.lower() for c in cells]
        link_col = next((j for j, c in enumerate(low) if c in _LINK_HEADERS), None)
        if link_col is None:
            continue
        acc_col = next((j for j, c in enumerate(low) if c in _ACCOUNT_HEADERS), None)
        cat_col = next((j for j, c in enumerate(low) if c in _CATEGORY_HEADERS), None)
        return i, {"link": link_col, "account": acc_col, "category": cat_col}
    return None


def _rows_from_grid(grid) -> list:
    grid = [cells for cells in grid if any(cells)]
    if not grid:
        return []

    header = _header_index(grid)
    rows = []
    if header:
        start, cmap = header
        for cells in grid[start + 1:]:
            link = cells[cmap["link"]] if cmap["link"] < len(cells) else ""
            if not _URL_RE.search(link):
                # some sheets put the URL in a different column than the header
                link = next((c for c in cells if _URL_RE.search(c)), "")
            if not link:
                continue
            account = cells[cmap["account"]] if cmap["account"] is not None \
                and cmap["account"] < len(cells) else ""
            category = cells[cmap["category"]] if cmap["category"] is not None \
                and cmap["category"] < len(cells) else ""
            rows.append(_row(category, account, link))
    else:
        # plain list: URL cell = link (account from a sibling text cell);
        # a URL-less line becomes the running category header.
        category = "Uncategorized"
        for cells in grid:
            if cells and cells[0].lstrip().startswith("#"):
                continue  # comment line
            url = next((c for c in cells if _URL_RE.search(c)), "")
            if url:
                account = next((c for c in cells
                                if c and not _URL_RE.search(c)
                                and not re.fullmatch(r"[\d,]+", c)), "")
                rows.append(_row(category, account, url))
            else:
                joined = " ".join(c for c in cells if c)
                if joined.lower() not in _IGNORE_HEADERS:
                    category = joined or category

    kept = [r for r in rows if is_x_url(r["link"])]
    dropped = len(rows) - len(kept)
    if dropped:
        print(f"[input] skipped {dropped} non-X link(s) — this tool is X-only")
    return kept


# --------------------------------------------------------------------------- #
# Dispatch
# --------------------------------------------------------------------------- #
def load(source: str) -> list:
    s = (source or "").strip()
    if s in ("-", "paste", "stdin"):
        return load_stdin()
    print(f"[input] reading {s}")
    return load_excel(s)   # .xlsx workbook; falls back to a plain link list if it's text
